"""
One-time script to load kitchen menu items from Excel into kitch_item_catalg.
Source: ~/Desktop/Rajashree Kitchen Items.xlsx
"""

import os
import sys
import openpyxl
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from dotenv import load_dotenv
load_dotenv()

import psycopg2

EXCEL_PATH = Path.home() / "Desktop" / "Rajashree Kitchen Items.xlsx"

# Category IDs (existing in DB)
CAT_MAIN_COURSE = 2
CAT_APPETIZERS = 1
CAT_DESSERTS = 3

# Descriptions keyed by normalized item name (lowercase, stripped)
DESCRIPTIONS = {
    "chilli paneer": "Crispy paneer tossed in spicy Indo-Chinese chilli sauce, made with home-made paneer.",
    "home made paneer": "Crispy paneer tossed in spicy Indo-Chinese chilli sauce, made with home-made paneer.",
    "chilli paneer \nhome made paneer": "Crispy home-made paneer tossed in a spicy Indo-Chinese chilli sauce.",
    "soya chilli": "Soya chunks stir-fried in a tangy and spicy chilli sauce.",
    "butter paneer": "Tender paneer simmered in a creamy, mildly spiced tomato-butter gravy.",
    "chana dal fry": "Yellow split lentils tempered with cumin, garlic, and aromatic spices.",
    "toor dal fry": "Split pigeon peas cooked with onion, tomato, and traditional Indian spices.",
    "aloo gobi mattar": "Potatoes, cauliflower, and peas cooked together with fragrant spices.",
    "navaratna korma": "Nine-vegetable korma in a rich, creamy nut-based sauce.",
    "dahi vada": "Soft lentil dumplings soaked in chilled yogurt and topped with chutneys.",
    "aloo dam": "Tender baby potatoes slow-cooked in a robust spiced tomato gravy.",
    "rajma curry": "Red kidney beans braised in a thick, spiced onion-tomato masala.",
    "chola curry": "Hearty chickpea curry cooked with tangy tamarind and warming spices.",
    "aloo parbal korma": "Potatoes and pointed gourd in a delicate, mildly spiced korma gravy.",
    "kofta curry": "Soft vegetable dumplings served in a rich and aromatic tomato-based curry.",
    "idli+sambar": "Steamed rice cakes served with piping hot lentil-vegetable sambar.",
    "vada+sambar": "Crispy lentil fritters paired with tangy, spiced lentil-vegetable sambar.",
    "veg manchurian": "Vegetable balls in a tangy, spiced Indo-Chinese Manchurian sauce.",
    "paratha(3)": "Three flaky whole-wheat flatbreads, served warm with butter.",
    "rice kheer": "Slow-cooked rice pudding with milk, sugar, cardamom, and nuts.",
    "vermesil kheer": "Roasted vermicelli cooked in sweetened milk with cardamom and dry fruits.",
    "white pasta": "Pasta in a creamy béchamel sauce with mild herbs.",
    "red pasta": "Pasta tossed in a spiced tomato-based arrabbiata sauce.",
    "veg biriyani": "Fragrant basmati rice layered with seasoned mixed vegetables and saffron.",
    "paneer mattar pulao": "Aromatic basmati rice cooked with paneer and green peas.",
    "samosa chat": "Flaky samosas topped with yogurt, chutneys, and crispy sev.",
    "crispy gobi fry": "Cauliflower florets battered and fried to a golden, crispy finish.",
    "gobi 65": "Deep-fried spiced cauliflower in the style of the classic 65 recipe.",
    "daal makhani": "Whole black lentils slow-cooked overnight in a buttery, creamy tomato sauce.",
    "bhindi masala": "Stir-fried okra with onions, tomatoes, and a blend of Indian spices.",
    "brinjal masala": "Tender eggplant cooked in a tangy, spice-rich masala sauce.",
    "kadai paneer": "Paneer cubes and peppers tossed in a bold, freshly ground kadai spice sauce.",
    "paneer shimla mirch": "Paneer and capsicum stir-fried in a vibrant, spiced onion-tomato gravy.",
    "veg noodles": "Stir-fried noodles with crisp vegetables in a light Indo-Chinese sauce.",
    "onion pakoda": "Crispy golden fritters made with sliced onions and spiced chickpea batter.",
    "spinach pakoda": "Fresh spinach leaves dipped in seasoned chickpea batter and fried crisp.",
    "palak paneer": "Fresh spinach puree with soft paneer cubes in a mildly spiced sauce.",
    "jeera rice": "Basmati rice tempered with cumin seeds, served fragrant and fluffy.",
    "basanti pulao": "Sweet yellow rice made with saffron, ghee, cashews, and raisins.",
    "dhoka paneer": "Spiced chickpea flour cakes cooked with paneer in a tangy tomato curry.",
    "lasuni palak": "Creamed spinach generously tempered with garlic and mild spices.",
    # Non-veg
    "mutton curry": "Slow-cooked tender mutton in a deeply spiced onion-tomato gravy.",
    "chicken curry": "Classic home-style chicken curry with aromatic whole spices.",
    "chilli chicken": "Crispy chicken tossed in a tangy, spicy Indo-Chinese chilli sauce.",
    "chicken lollipop(5)": "Five crispy chicken lollipops marinated and fried with Indian spices.",
    "chicken manchurian": "Crispy chicken balls in a bold, tangy Indo-Chinese Manchurian sauce.",
    "dragon chicken": "Crispy fried chicken strips tossed in a fiery dragon sauce.",
    "chicken biriyani": "Fragrant basmati rice slow-cooked with spiced chicken and saffron.",
    "mutton biriyani": "Slow-cooked basmati rice layered with tender, spiced mutton.",
    "shrimp biriyani": "Aromatic basmati rice cooked with juicy shrimp and fragrant spices.",
    "shrimp popcorn": "Bite-sized crispy shrimp seasoned with a zesty spice blend.",
    "chilli shrimp": "Succulent shrimp tossed in a bold, spicy Indo-Chinese chilli sauce.",
    "shrimp curry": "Plump shrimp simmered in a tangy, coconut-spiced curry sauce.",
    "egg labadar": "Boiled eggs cooked in a rich, spiced onion-tomato-cashew gravy.",
    "egg curry": "Hard-boiled eggs in a hearty, home-style spiced tomato gravy.",
    "egg biriyani": "Fragrant basmati rice layered with spiced boiled eggs.",
    "chicken egg roll": "Flaky paratha rolled with spiced scrambled egg and chicken filling.",
    "chicken 65": "Deep-fried spiced chicken in the iconic South Indian 65 style.",
    "butter chicken": "Tender chicken in a velvety, mildly spiced tomato-butter-cream sauce.",
    "kadai chicken": "Chicken and peppers cooked in a bold freshly ground kadai spice blend.",
    "fish curry": "Fish fillets simmered in a tangy, coconut-spiced South Indian curry.",
    # Snacks
    "samosa(3)": "Three crispy pastry pockets filled with spiced potatoes and peas.",
    "veg puff(2)": "Two flaky pastry puffs filled with a savory spiced vegetable mix.",
    "egg puff(2)": "Two golden pastry puffs filled with a seasoned boiled egg.",
    "veg cutlet(3)": "Three pan-fried vegetable patties coated in crispy breadcrumbs.",
    "rasagola": "Soft, spongy cottage cheese balls soaked in light sugar syrup.",
    # Sweets
    "jelebi(1lb)": "One pound of freshly fried crispy spirals soaked in sugar syrup.",
    "chamcham": "Soft Bengali milk sweets dipped in sugar syrup and dusted with coconut.",
    "kajukatli(3)": "Three diamond-shaped cashew fudge bites with a silver leaf garnish.",
    "khaja": "Flaky, layered deep-fried sweet pastry soaked in sugar syrup.",
    "rasmalai(4)": "Four soft cottage cheese patties soaked in saffron-cardamom-flavored cream.",
    "rasabali(4)": "Four fried cottage cheese patties soaked in sweet thickened milk.",
    "gujia": "Deep-fried pastry stuffed with a sweet khoya and dry fruit filling.",
    "nan khatai": "Buttery, crumbly Indian shortbread biscuits with cardamom flavor.",
    "coconut sweet (6)": "Six bite-sized coconut laddoos made with fresh coconut and jaggery.",
    "besan ladoo (6)": "Six round gram flour sweets roasted in ghee and rolled with cardamom.",
    "rasagulla(10)": "Ten soft, spongy paneer balls soaked in chilled sugar syrup.",
    "bundi ladoo ( 1lb) ": "One pound of golden chickpea flour pearls formed into sweet laddoos.",
    "bundi ladoo(1lb) ": "One pound of golden chickpea flour pearls formed into sweet laddoos.",
    "gajar halwa": "Classic North Indian carrot pudding slow-cooked with milk, ghee, and nuts.",
    "gulab jamun": "Soft milk-solid dumplings soaked in rose-flavored sugar syrup.",
}

def get_description(name):
    key = name.strip().lower()
    if key in DESCRIPTIONS:
        return DESCRIPTIONS[key]
    # Fallback: generic description based on key terms
    if any(k in key for k in ['paneer', 'aloo', 'gobi', 'dal', 'daal', 'palak', 'rajma', 'chola', 'veg']):
        return f"Delicious vegetarian dish - {name.strip().title()}."
    if any(k in key for k in ['chicken', 'mutton', 'shrimp', 'egg', 'fish']):
        return f"Flavorful non-vegetarian specialty - {name.strip().title()}."
    if any(k in key for k in ['kheer', 'halwa', 'ladoo', 'jelebi', 'rasagol', 'barfi', 'gulab', 'rasmalai']):
        return f"Traditional Indian sweet - {name.strip().title()}."
    return f"A signature Rajashree Kitchen specialty - {name.strip().title()}."


def parse_price(value):
    """Return float price or None if not numeric."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    # String price like "$1.50 each" — not a numeric per-plate price
    return None


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws_plate = wb['PerPlatePrice']
    ws_party = wb['New_Party_order_price']

    # Collect party item names for confirming is_catering
    party_names = set()
    for row in ws_party.iter_rows(min_row=2, values_only=True):
        # Veg names in col B (idx 1), nonveg in col H (idx 7)
        for idx in [1, 7]:
            if row[idx] is not None:
                party_names.add(str(row[idx]).strip().lower())

    # Parse PerPlatePrice
    items = []
    for row in ws_plate.iter_rows(min_row=2, values_only=True):
        # Main course veg: cols B (idx 1), C (idx 2)
        if row[1] is not None:
            name = str(row[1]).strip()
            price = parse_price(row[2])
            in_party = name.lower() in party_names
            items.append({
                'name': name,
                'price': price,
                'category_id': CAT_MAIN_COURSE,
                'is_catering': in_party,
                'description': get_description(name),
            })

        # Main course nonveg: cols F (idx 5), G (idx 6)
        if row[5] is not None:
            name = str(row[5]).strip()
            price = parse_price(row[6])
            in_party = name.lower() in party_names
            items.append({
                'name': name,
                'price': price,
                'category_id': CAT_MAIN_COURSE,
                'is_catering': in_party,
                'description': get_description(name),
            })

        # Snacks: cols J (idx 9), K (idx 10)
        if row[9] is not None:
            name = str(row[9]).strip()
            price = parse_price(row[10])
            items.append({
                'name': name,
                'price': price,
                'category_id': CAT_APPETIZERS,
                'is_catering': False,
                'description': get_description(name),
            })

        # Sweets: cols N (idx 13), O (idx 14)
        if row[13] is not None:
            name = str(row[13]).strip()
            price = parse_price(row[14])
            items.append({
                'name': name,
                'price': price,
                'category_id': CAT_DESSERTS,
                'is_catering': False,
                'description': get_description(name),
            })

    # Connect to DB
    conn = psycopg2.connect(
        host=os.getenv('DB_HOST', 'localhost'),
        port=os.getenv('DB_PORT', 5432),
        dbname=os.getenv('DB_NAME', 'kitchen_db'),
        user=os.getenv('DB_USER', 'kitchen_user'),
        password=os.getenv('DB_PASSWORD'),
    )
    cur = conn.cursor()

    # Fetch existing item names to skip duplicates
    cur.execute("SELECT LOWER(kic_name) FROM kitch_item_catalg")
    existing = {r[0] for r in cur.fetchall()}

    inserted = 0
    skipped = 0
    no_price = 0
    for item in items:
        if item['name'].lower() in existing:
            print(f"  SKIP (exists): {item['name']}")
            skipped += 1
            continue

        if item['price'] is None:
            print(f"  SKIP (no price): {item['name']}")
            no_price += 1
            continue

        cur.execute(
            """
            INSERT INTO kitch_item_catalg (kic_name, kic_price, category_id, description, is_catering, is_active)
            VALUES (%s, %s, %s, %s, %s, TRUE)
            """,
            (item['name'], item['price'], item['category_id'], item['description'], item['is_catering'])
        )
        existing.add(item['name'].lower())  # avoid duplicate within this run
        print(f"  INSERT: {item['name']} | price={item['price']} | catering={item['is_catering']}")
        inserted += 1

    conn.commit()
    cur.close()
    conn.close()

    print(f"\nDone. Inserted: {inserted}, Skipped (exists): {skipped}, Skipped (no price): {no_price}")


if __name__ == '__main__':
    main()
